"""
Excel工具类
"""

from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.workbook import Workbook

import Utils.Constant as Constant
import Utils.DataUtils as Data
from Utils.FileUtils import FileOper


# 定义自定义异常类
class CustomError(Exception):
    def __init__(self, message):
        self.message = message


def apply_format(cell, is_header):
    font = Font(name="LXGWWenKaiGBScreen", size=14, color="000000", bold=False) if not is_header else Font(
        name="LXGWWenKaiGBScreen", size=18, color="ffffff")
    alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, indent=0)
    fill = PatternFill(patternType="solid", fgColor="ffffcc", bgColor="ff2600") if not is_header else PatternFill(
        patternType="solid", fgColor="5e7ce0", bgColor="9bc2e6")
    side = Side(style="thin", color="000000")
    border = Border(top=side, bottom=side, left=side, right=side, diagonal=side)

    cell.font = font
    cell.alignment = alignment
    cell.fill = fill
    cell.border = border


def export_artifact_data(data_list):
    """
    导出背包圣遗物数据
    :param data_list 数据列表，格式：[[...],[...]]
    """
    # 定义表头键值对列表
    head_map = {
        "main_name": "套装",
        "children_name": "子件名称",
        "slot": "部位",
        "equip_role": "装备角色",
        "main_tag_name": "主词条",
        "main_tag_value": "主词条属性值",
        "children_tag_name_1": "词条1",
        "children_tag_value_1": "属性值1",
        "children_tag_name_2": "词条2",
        "children_tag_value_2": "属性值2",
        "children_tag_name_3": "词条3",
        "children_tag_value_3": "属性值3",
        "children_tag_name_4": "词条4",
        "children_tag_value_4": "属性值4"
    }

    # 创建一个工作簿
    wb = Workbook()
    # 获取当前活动的工作表
    ws = wb.active
    # 设置工作表的名称
    ws.title = u'圣遗物数据'

    # 设置表头
    for col, header in enumerate(head_map.values(), start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header

    # 设置表头的列宽
    for cell in ws[1]:
        ws.column_dimensions[cell.column_letter].width = max(len(str(cell.value)) + 20, 20)

    # 根据字段映射将数据插入到相应的单元格中
    for row, row_data in enumerate(data_list, start=2):
        for field, col in head_map.items():
            cell = ws.cell(row=row, column=list(head_map.keys()).index(field) + 1)
            cell.value = row_data.get(field, "--")

    for row in ws:
        for cell in row:
            is_header = cell.row == 1
            apply_format(cell, is_header)

    config_dir = Data.settings.value("config_dir")
    if not config_dir:
        raise CustomError("未设置本地数据目录")

    target_dir = FileOper.get_dir(f"{config_dir}/{Constant.out_dir}")

    # 将工作簿保存到磁盘
    wb.save(f'{target_dir}/圣遗物数据.xlsx')
    FileOper.open_dir(target_dir)
